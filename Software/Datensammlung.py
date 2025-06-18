from openpyxl import load_workbook

workbook=None

def open_book(name):
    workbook = load_workbook('meinedatei.xlsx')


def new_bookpage(name):
    if name not in workbook.sheetnames:
        workbook.create_sheet(title=name)

def switch_active_page(name):
    workbook.active = workbook[name]

def save_data():
    workbook.save('datensammlung.xlsx')
