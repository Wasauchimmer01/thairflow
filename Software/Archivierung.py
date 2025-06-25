from openpyxl import load_workbook
from test import write_data_xlsx


workbook = load_workbook('Daten.xlsx')

def new_bookpage(name):
    if name not in workbook.sheetnames:
        workbook.create_sheet(title=name)

def switch_active_page(name):
    workbook.active = workbook[name]

def save_data():
    workbook.save('Daten.xlsx')

def write_data(data,page):
    switch_active_page(page)
    write_data_xlsx(data)
