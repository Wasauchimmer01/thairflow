from datetime import datetime, timedelta
import csv
import time
import subprocess
from openpyxl import Workbook, load_workbook
from software.formaldehyd import SFA30Sensor
from software.motionsensor import MotionSensor
from software.differenzdruck_810 import SdpSensor
from software.partikelsensor import Sen66Sensor
from software.gyroskop_fenster import GyroSensor
#import csv
import os
import shutil
from software.emailer import send_report

SENSOR_MAP = {
    "5D": SFA30Sensor,
    "5E": SFA30Sensor,  # angenommen 0x5e
    "48": MotionSensor,
    "49": MotionSensor,
    "25": SdpSensor,
    "26": SdpSensor,
    "68": GyroSensor,
    "69": GyroSensor,
    "6B": Sen66Sensor,
}

def get_available_sensors():
    bus_num = '1'
    output = subprocess.check_output(['i2cdetect', '-y', bus_num], text=True)
    devices = []
    for line in output.splitlines():
        if ':' not in line:
            continue  # Überspringe Kopfzeile
        parts = line.split(':')[1].split()
        for p in parts:
            if p != '--':
                devices.append(p.upper())
    print(f"i2c-{bus_num}: " + ", ".join(f"0x{addr}" for addr in devices))

    available_sensors = []
    for addr in devices:
        if addr in SENSOR_MAP:
            sensor_class = SENSOR_MAP[addr]
            slave_addr = int(addr, 16)
            sensor = sensor_class(slave_addr)
            available_sensors.append(sensor)

    return available_sensors

def update_log(sensor, values, log):
    timestamp = datetime.now().isoformat()
    for name, value in zip(sensor.values, values):
        log[sensor.name+ "_" + str(sensor.address)][name]["data"].append(value)
        log[sensor.name+ "_" + str(sensor.address)][name]["time"].append(timestamp)
    return log

def print_data_log(data_log):
    for sensor_name, measurements in data_log.items():
        print(f"Sensor: {sensor_name}")
        for value_name, details in measurements.items():
            print(f"  {value_name} ({details['unit']}):")
            for t, v in zip(details["time"], details["data"]):
                print(f"    {t}: {v}")
        print("-" * 40)

def reset_log(data_log):
    for sensor_name in data_log:
        for value_name in data_log[sensor_name]:
            data_log[sensor_name][value_name]["data"].clear()
            data_log[sensor_name][value_name]["time"].clear()
    return data_log



def make_xlsx(data_log, filename="daten"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sensor Data"
    
    # Build headers
    sensor_row = []
    value_row = []
    unit_row = []
    
    # First column reserved for timestamps
    sensor_row.append("")
    value_row.append("")
    unit_row.append("Time")
    
    for sensor_name, measurements in data_log.items():
        for value_name, details in measurements.items():
            sensor_row.append(sensor_name)
            value_row.append(value_name)
            unit_row.append(details["unit"])
    
    ws.append(sensor_row)
    ws.append(value_row)
    ws.append(unit_row)
    
    full_filename = filename+"_"+ str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
    wb.save(full_filename)
    print(f"Data saved to {full_filename}")
    return full_filename


def append_measurements_to_xlsx(data_log, filename):
    wb = load_workbook(filename)
    ws = wb.active

    # Get timestamp from the first available measurement
    first_sensor = next(iter(data_log.values()))
    first_value = next(iter(first_sensor.values()))
    if not first_value["time"]:
        print("No new data to append.")
        return
    timestamp = first_value["time"][0]

    # Build row: timestamp + all measurement values in the same column order as header
    row = [timestamp]
    for sensor_name, measurements in data_log.items():
        for value_name, details in measurements.items():
            row.append(details["data"][0] if details["data"] else "")

    ws.append(row)
    wb.save(filename)
    print(f"Data appended to {filename}")

def make_csv(data_log, filename="daten"):
    rows = []
    
    # Build headers
    sensor_row = []
    value_row = []
    unit_row = []

    # First column reserved for timestamps
    sensor_row.append("")
    value_row.append("")
    unit_row.append("Time")

    for sensor_name, measurements in data_log.items():
        for value_name, details in measurements.items():
            sensor_row.append(sensor_name)
            value_row.append(value_name)
            unit_row.append(details["unit"])

    rows.append(sensor_row)
    rows.append(value_row)
    rows.append(unit_row)

    # Optional: hier könntest du auch Messwerte hinzufügen, falls vorhanden
    # z.B. rows.append(["12:00", val1, val2, ...])

    full_filename = filename + "_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".csv"
    with open(full_filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(rows)

    print(f"Data saved to {full_filename}")
    return full_filename

def append_measurements_to_csv(data_log, filename):
    # Falls Dateiname noch kein ".csv" hat
    if not filename.endswith(".csv"):
        filename += ".csv"

    # Get timestamp
    first_sensor = next(iter(data_log.values()))
    first_value = next(iter(first_sensor.values()))
    if not first_value["time"]:
        print("No new data to append.")
        return
    timestamp = first_value["time"][0]

    # Baue Datenzeile im selben Format wie in der Kopfzeile
    row = [timestamp]
    for sensor_name, measurements in data_log.items():
        for value_name, details in measurements.items():
            row.append(details["data"][0] if details["data"] else "")

    # An die CSV anhängen
    with open(filename, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(row)

    print(f"Data appended to {filename}")

def make_backup_file(time, backup_intervall, filename):
    delta_t = timedelta(hours=backup_intervall)
    if datetime.now() - time < delta_t:
        return time
    else:
        basename, extension = os.path.splitext(filename)
        b_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        copy_name = f"{basename}_backup_{b_time}.{extension}"
        shutil.copy(filename, copy_name)
        return datetime.now()


if __name__ == "__main__":
    sensors = get_available_sensors()
    data_log = {}
    backup_intervall = 1

    # make data log dict
    for sensor in sensors:
        print(sensor.name, sensor.address)
    data_log = {
        sensor.name+ "_" + str(sensor.address): {
            value_name: {"unit": unit, "data": [], "time": []}
            for value_name, unit in zip(sensor.values, sensor.units)
        }
        for sensor in sensors
    }
    filename_xl = make_xlsx(data_log)
    #filename_csv = make_csv(data_log)
    filename_csv = "daten.csv"
    
    # Initialize timers
    last_backup_time = datetime.now()
    backup_intervall_hours = 1
    last_email_time = datetime.now() - timedelta(minutes=59)

    print("Measurement starts")
    save_time = datetime.now()
    while True:
        try:
            # used as timer to trigger reading every secound
            now = datetime.now()
            for sensor in sensors:
                try:
                    data = sensor.read_measurements()
                    # could use error handler
                    data_log = update_log(sensor, data, data_log)
                except Exception as e:
                    print(f"Error reading from sensor {sensor.name}: {e}")
            #print_data_log(data_log)
            append_measurements_to_xlsx(data_log, filename_xl)
            #append_measurements_to_csv(data_log, filename_csv)
            append_measurements_to_csv(data_log, filename=filename_csv)  # EMAIL ADDED: keep a rolling daten.csv
            data_log = reset_log(data_log)
            
            # Hourly backup
            last_backup_time = make_backup_file(
                last_backup_time,
                backup_interval_hours,
                filename_xl
            )
             # ── SEND CSV BY E-MAIL EVERY 59 MINUTES ─────────────────────────
            if datetime.now() - last_email_time >= timedelta(minutes=59):
                try:
                    send_report(
                        filepath=filename_csv,
                        subject=f"Sensor Log {now:%Y-%m-%d %H:%M}",
                        body="Attached is the latest sensor CSV log."
                    )
                    last_email_time = datetime.now()
                    print(f"[Email] Sent daten.csv at {datetime.now():%H:%M}")
                except Exception as e:
                    print(f"[Email] Failed to send: {e}")
            # save_time = make_backup_file(save_time,backup_intervall, filename_xl) experiment make sure that is 
            time.sleep(1 - now.microsecond / 1_000_000)
        except KeyboardInterrupt:
            print("Measurement stopped by user.")
            break
         
    
    for sensor in sensors:
        sensor.stop_sensor()
        print(f"{sensor.name} stopped.")